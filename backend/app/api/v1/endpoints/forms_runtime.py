from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timedelta, timezone
import re
from typing import Any
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response as FastApiResponse
from sqlalchemy import and_, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import (
    AccessControl,
    AppUser,
    Form,
    FormElement,
    FormPage,
    Response,
    ResponseAnswer,
    UploadedFile,
)
from app.schemas import (
    BuilderElementOut,
    FormDraftResponse,
    FormDraftSaveRequest,
    FormStoredResponse,
    FormStoredResponsesResponse,
    FormSubmitAnswersRequest,
    FormSubmitAnswersResponse,
    PublicFormDetailResponse,
)
from app.security.auth_dependencies import get_current_user, get_optional_user
from app.services.forms_mapping import build_form_detail_response

router = APIRouter(prefix="/forms", tags=["forms"])

ANON_EMAIL = "anonymous@etu-forms.local"
ANON_NAME = "Anonymous Respondent"
SNILS_DIGITS_ONLY_RE = re.compile(r"\D")
SNILS_TRIPLE_DIGIT_RE = re.compile(r"(\d)\1\1")
SNILS_FORMATTED_RE = re.compile(r"^\d{3}-\d{3}-\d{3}\s\d{2}$")
SNILS_PLAIN_RE = re.compile(r"^\d{11}$")
SNILS_CHECKSUM_THRESHOLD = 1_001_998
SNILS_INVALID_MSG = "Invalid SNILS"
SNILS_REPEATED_DIGITS_MSG = "Invalid SNILS repeated digits"
SNILS_CHECKSUM_MSG = "Invalid SNILS checksum"
EXPORT_HEADERS = {
    "ru": {
        "responder_name": "Респондент",
        "responder_email": "Email",
        "status": "Статус",
        "created_at": "Создан",
        "completed_at": "Завершён",
        "version": "Версия формы",
        "other": "Другое",
    },
    "en": {
        "responder_name": "Responder",
        "responder_email": "Email",
        "status": "Status",
        "created_at": "Created at",
        "completed_at": "Completed at",
        "version": "Form version",
        "other": "Other",
    },
}

EXPORT_LIST_SEP_CSV = " ; "

EXPORT_SHEETS = {
    "ru": {"responses": "Ответы", "summary": "Итоги", "statistics": "Статистика"},
    "en": {"responses": "Responses", "summary": "Summary", "statistics": "Statistics"},
}

EXPORT_SUMMARY = {
    "ru": {
        "metric": "Показатель",
        "value": "Значение",
        "form_title": "Название формы",
        "exported_at": "Дата и время выгрузки",
        "responses_in_export": "Ответов в выгрузке",
        "status_submitted": "Отправлено",
        "status_cancelled": "Отменено",
        "form_version": "Версия формы",
    },
    "en": {
        "metric": "Metric",
        "value": "Value",
        "form_title": "Form title",
        "exported_at": "Export date and time",
        "responses_in_export": "Responses in export",
        "status_submitted": "Submitted",
        "status_cancelled": "Cancelled",
        "form_version": "Form version",
    },
}

EXPORT_STATS = {
    "ru": {
        "dimension": "Показатель",
        "value_col": "Значение",
        "nonempty": "Ответов с данными по полю",
        "min": "Минимум",
        "max": "Максимум",
        "average": "Среднее",
        "distribution": "Распределение",
    },
    "en": {
        "dimension": "Metric",
        "value_col": "Value",
        "nonempty": "Non-empty answers for field",
        "min": "Minimum",
        "max": "Maximum",
        "average": "Average",
        "distribution": "Distribution",
    },
}


def _enum_value(x: Any) -> Any:
    return x.value if hasattr(x, "value") else x

def _access_not_expired():
    now = datetime.utcnow()
    return and_(
        or_(AccessControl.starts_at.is_(None), AccessControl.starts_at <= now),
        or_(AccessControl.expires_at.is_(None), AccessControl.expires_at > now),
    )


def _protected_link_key(form: Form) -> str | None:
    settings = form.settings_json if isinstance(form.settings_json, dict) else {}
    raw = settings.get("privateLinkKey")
    return raw if isinstance(raw, str) and raw else None


def _public_settings(form: Form) -> dict[str, Any] | None:
    if not isinstance(form.settings_json, dict):
        return None
    settings = dict(form.settings_json)
    settings.pop("privateLinkKey", None)
    return settings or None


def _increment_link_views(form: Form) -> None:
    settings = dict(form.settings_json) if isinstance(form.settings_json, dict) else {}
    try:
        current = int(settings.get("linkViews", 0))
    except (TypeError, ValueError):
        current = 0
    settings["linkViews"] = max(0, current) + 1
    form.settings_json = settings


def _is_auth_required(form: Form) -> bool:
    return _enum_value(form.access_mode) == "private"


def _now_for_compare(form: Form) -> datetime:
    # Keep timezone compatibility with DB value (aware/naive).
    reference = form.start_at or form.end_at
    if reference is not None and reference.tzinfo is not None:
        return datetime.now(reference.tzinfo)
    return datetime.utcnow()


def _check_form_access(form: Form, key: str | None, current_user: AppUser | None) -> None:
    status_value = _enum_value(form.status)
    if status_value != "submitted":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")

    if _is_auth_required(form) and current_user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")

    now = _now_for_compare(form)
    if form.start_at is not None and form.start_at > now:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Form is not open yet")
    if form.end_at is not None and form.end_at < now:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Form is closed")

    access_mode = _enum_value(form.access_mode)
    if access_mode in {"private", "unauthenticated"}:
        expected = _protected_link_key(form)
        provided = key.strip() if isinstance(key, str) else None
        if not expected or provided != expected:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Invalid link key")


async def _get_published_form(db: AsyncSession, form_id: int) -> Form:
    result = await db.execute(select(Form).where(Form.form_id == form_id))
    form = result.scalar_one_or_none()
    if not form or form.status == "deleted":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")
    return form


async def _resolve_latest_submitted_form(db: AsyncSession, form: Form) -> Form:
    current = form
    visited = {current.form_id}

    while True:
        result = await db.execute(
            select(Form)
            .where(Form.prev_form_id == current.form_id)
            .where(Form.status == "submitted")
            .order_by(Form.version.desc(), Form.updated_at.desc(), Form.form_id.desc())
            .limit(1)
        )
        next_form = result.scalar_one_or_none()
        if next_form is None:
            break
        if next_form.form_id in visited:
            break
        visited.add(next_form.form_id)
        current = next_form

    return current


async def _get_or_create_anonymous_user(db: AsyncSession) -> AppUser:
    result = await db.execute(select(AppUser).where(AppUser.email == ANON_EMAIL))
    user = result.scalar_one_or_none()
    if user:
        return user

    user = AppUser(
        name=ANON_NAME,
        email=ANON_EMAIL,
        etu_id=None,
        is_admin=False,
    )
    db.add(user)
    await db.flush()
    return user


async def _resolve_responder(
    db: AsyncSession,
    current_user: AppUser | None,
) -> AppUser:
    if current_user is not None:
        return current_user
    return await _get_or_create_anonymous_user(db)


def _extract_file_ids(value: Any) -> list[int]:
    if not isinstance(value, dict):
        return []
    raw_ids = value.get("file_ids")
    if not isinstance(raw_ids, list):
        return []

    normalized: list[int] = []
    for item in raw_ids:
        try:
            file_id = int(item)
        except (TypeError, ValueError):
            continue
        if file_id > 0:
            normalized.append(file_id)
    unique_ids = list(dict.fromkeys(normalized))
    return unique_ids[:10]


def _apply_answer_value(answer_row: ResponseAnswer, value: Any) -> None:
    answer_row.value_text = None
    answer_row.value_number = None
    answer_row.value_bool = None
    answer_row.value_date = None
    answer_row.value_time = None
    answer_row.value_json = None

    if value is None:
        return

    answer_row.value_json = value
    if isinstance(value, bool):
        answer_row.value_bool = value
    elif isinstance(value, int):
        answer_row.value_number = value
    elif isinstance(value, str):
        answer_row.value_text = value


def _snils_checksum(number: str) -> int:
    total = sum(int(digit) * weight for weight, digit in enumerate(reversed(number), start=1))
    if total < 100:
        return total
    if total in (100, 101):
        return 0

    remainder = total % 101
    if remainder in (100, 101):
        return 0
    return remainder


def _validate_snils_value(value: Any) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        return SNILS_INVALID_MSG

    candidate = value.strip()
    if candidate == "":
        return None

    if not (SNILS_PLAIN_RE.fullmatch(candidate) or SNILS_FORMATTED_RE.fullmatch(candidate)):
        return "SNILS must match XXX-XXX-XXX YY or contain 11 digits"

    digits = SNILS_DIGITS_ONLY_RE.sub("", candidate)
    number = digits[:9]
    checksum = digits[9:]

    if SNILS_TRIPLE_DIGIT_RE.search(number):
        return SNILS_REPEATED_DIGITS_MSG

    if int(number) > SNILS_CHECKSUM_THRESHOLD:
        expected = _snils_checksum(number)
        if int(checksum) != expected:
            return SNILS_CHECKSUM_MSG
    return None


def _validate_semantic_answer(element: FormElement, value: Any) -> str | None:
    semantic = _enum_value(element.semantic)
    if semantic == "snils":
        return _validate_snils_value(value)
    return None


def _file_url(file_id: int, token: str) -> str:
    return f"/api/v1/files/{file_id}/download?token={token}"


def _deserialize_answer_value(
    answer_row: ResponseAnswer,
    attachments: list[dict[str, Any]],
) -> Any:
    if attachments:
        return attachments
    if answer_row.value_json is not None:
        return answer_row.value_json
    if answer_row.value_bool is not None:
        return answer_row.value_bool
    if answer_row.value_number is not None:
        return answer_row.value_number
    if answer_row.value_text is not None:
        return answer_row.value_text
    if answer_row.value_date is not None:
        return answer_row.value_date.isoformat()
    if answer_row.value_time is not None:
        return answer_row.value_time.isoformat()
    return None


async def _ensure_editor_or_owner(
    db: AsyncSession,
    form_id: int,
    current_user: AppUser,
    allowed_roles: tuple[str, ...] = ("editor",),
) -> Form:
    result = await db.execute(select(Form).where(Form.form_id == form_id))
    form = result.scalar_one_or_none()
    if not form or form.status == "deleted":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")

    if form.user_id == current_user.user_id:
        return form

    if not allowed_roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    access = await db.execute(
        select(AccessControl)
        .where(AccessControl.form_id == form_id)
        .where(AccessControl.user_id == current_user.user_id)
        .where(AccessControl.role.in_(allowed_roles))
        .where(_access_not_expired())
    )
    if not access.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    return form


def _sorted_form_elements(
    elements: list[FormElement],
    page_index_by_id: dict[int, int],
) -> list[FormElement]:
    return sorted(
        elements,
        key=lambda item: (
            page_index_by_id.get(item.page_id, 10_000),
            item.position if item.position is not None else 10_000,
            item.element_id,
        ),
    )


def _extract_element_client_id(element: FormElement) -> str:
    settings = element.other_settings if isinstance(element.other_settings, dict) else {}
    client_id = settings.get("client_id")
    return str(client_id) if client_id is not None else str(element.element_id)


def _extract_export_answer_parts(value: Any, locale: str, list_sep: str) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            text = _stringify_export_value(item, locale, list_sep)
            if text:
                out.append(text)
        return out
    if isinstance(value, dict):
        if "attachments" in value and isinstance(value.get("attachments"), list):
            return _extract_export_answer_parts(value.get("attachments"), locale, list_sep)
        if "name" in value and isinstance(value.get("name"), str):
            name = str(value.get("name")).strip()
            return [name] if name else []
        if "url" in value and isinstance(value.get("url"), str):
            url = str(value.get("url")).strip()
            return [url] if url else []
        if "selected" in value or "otherSelected" in value:
            selected = value.get("selected")
            selected_parts: list[str] = []
            if isinstance(selected, list):
                selected_parts.extend([str(item) for item in selected if str(item).strip()])
            elif selected is not None and str(selected).strip():
                selected_parts.append(str(selected))
            if value.get("otherSelected"):
                other_text = str(value.get("otherText") or "").strip()
                other_label = EXPORT_HEADERS.get(locale, EXPORT_HEADERS["en"])["other"]
                selected_parts.append(f"{other_label}: {other_text}" if other_text else other_label)
            return selected_parts
        if "date" in value or "time" in value:
            date_part = str(value.get("date") or "").strip()
            time_part = str(value.get("time") or "").strip()
            joined = " ".join([part for part in [date_part, time_part] if part]).strip()
            return [joined] if joined else []
    text = _stringify_export_value(value, locale, list_sep).strip()
    return [text] if text else []


def _stringify_export_value(value: Any, locale: str, list_sep: str) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return list_sep.join(_extract_export_answer_parts(value, locale, list_sep))
    if isinstance(value, dict):
        parts = _extract_export_answer_parts(value, locale, list_sep)
        if parts:
            return list_sep.join(parts)
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _is_nonempty_export_answer(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, list):
        return len(value) > 0
    if isinstance(value, bool):
        return True
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, dict):
        return True
    return True


def _collect_numeric_values_for_stats(values: list[Any]) -> list[float] | None:
    nums: list[float] = []
    for raw in values:
        if raw is None:
            continue
        if isinstance(raw, bool):
            return None
        if isinstance(raw, (int, float)):
            nums.append(float(raw))
            continue
        if isinstance(raw, str):
            s = raw.strip().replace(",", ".")
            if not s:
                continue
            try:
                nums.append(float(s))
            except ValueError:
                return None
            continue
        return None
    return nums if nums else None


def _build_answer_map_by_response(
    answer_rows: list[ResponseAnswer],
    files_by_answer_id: dict[int, list[dict[str, Any]]],
    sorted_elements: list[FormElement],
) -> dict[int, dict[str, Any]]:
    element_client_map = {
        element.element_id: _extract_element_client_id(element) for element in sorted_elements
    }
    answer_rows_by_response_id: dict[int, dict[str, Any]] = {}
    for answer_row in answer_rows:
        client_id = element_client_map.get(answer_row.element_id, str(answer_row.element_id))
        attachments = files_by_answer_id.get(answer_row.answer_id, [])
        value = _deserialize_answer_value(answer_row, attachments)
        answer_rows_by_response_id.setdefault(answer_row.response_id, {})[client_id] = value
    return answer_rows_by_response_id


def _build_summary_export_rows(
    *,
    locale: str,
    form_title: str,
    form_version: int,
    responses: list[Response],
    export_at: datetime,
) -> list[list[str]]:
    labels = EXPORT_SUMMARY.get(locale, EXPORT_SUMMARY["en"])
    submitted = sum(1 for r in responses if _enum_value(r.status) == "submitted")
    cancelled = sum(1 for r in responses if _enum_value(r.status) == "cancelled")
    ts = export_at.isoformat(timespec="seconds")
    return [
        [labels["metric"], labels["value"]],
        [labels["form_title"], form_title],
        [labels["form_version"], str(form_version)],
        [labels["exported_at"], ts],
        [labels["responses_in_export"], str(len(responses))],
        [labels["status_submitted"], str(submitted)],
        [labels["status_cancelled"], str(cancelled)],
    ]


def _build_statistics_export_rows(
    *,
    locale: str,
    sorted_elements: list[FormElement],
    answer_rows_by_response_id: dict[int, dict[str, Any]],
    responses: list[Response],
) -> list[list[str]]:
    stats_labels = EXPORT_STATS.get(locale, EXPORT_STATS["en"])
    rows: list[list[str]] = [
        [stats_labels["dimension"], stats_labels["value_col"]],
    ]
    list_sep = EXPORT_LIST_SEP_CSV

    for element in sorted_elements:
        widget = _enum_value(element.widget)
        if widget in ("heading", "static_text"):
            continue
        client_id = _extract_element_client_id(element)
        label = (element.label or "").strip() or f"Field {element.element_id}"
        rows.append([f"— {label} —", ""])

        per_response: list[Any] = []
        for response_row in responses:
            raw = answer_rows_by_response_id.get(response_row.response_id, {}).get(client_id)
            per_response.append(raw)

        nonempty = [v for v in per_response if _is_nonempty_export_answer(v)]
        rows.append([stats_labels["nonempty"], str(len(nonempty))])

        if not nonempty:
            rows.append(["", ""])
            continue

        nums = _collect_numeric_values_for_stats(nonempty)
        if (
            nums is not None
            and len(nums) == len(nonempty)
            and widget in ("number_input", "rating")
        ):
            rows.append([stats_labels["min"], str(min(nums))])
            rows.append([stats_labels["max"], str(max(nums))])
            rows.append([stats_labels["average"], str(round(sum(nums) / len(nums), 6))])
            rows.append(["", ""])
            continue

        counter: dict[str, int] = {}
        for v in nonempty:
            for part in _extract_export_answer_parts(v, locale, list_sep):
                p = part.strip()
                if not p:
                    continue
                counter[p] = counter.get(p, 0) + 1
        if not counter:
            rows.append(["", ""])
            continue
        rows.append([stats_labels["distribution"], ""])
        for key in sorted(counter.keys(), key=lambda s: (-counter[s], s)):
            rows.append([key, str(counter[key])])
        rows.append(["", ""])

    while rows and rows[-1] == ["", ""]:
        rows.pop()
    return rows


def _sanitize_filename_part(value: str) -> str:
    # Windows/macOS-safe filename cleanup.
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", value or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip().strip(".")
    return cleaned


def _build_export_rows(
    *,
    form_version: int,
    responses: list[Response],
    answer_rows_by_response_id: dict[int, dict[str, Any]],
    users_map: dict[int, AppUser],
    sorted_elements: list[FormElement],
    locale: str,
    list_sep: str,
) -> tuple[list[str], list[list[str]]]:
    element_client_ids = [_extract_element_client_id(element) for element in sorted_elements]
    element_label_by_client_id = {
        _extract_element_client_id(element): (element.label or f"Field {element.element_id}").strip()
        or f"Field {element.element_id}"
        for element in sorted_elements
    }

    labels = EXPORT_HEADERS.get(locale, EXPORT_HEADERS["en"])
    headers = [
        labels["responder_name"],
        labels["responder_email"],
        labels["status"],
        labels["created_at"],
        labels["completed_at"],
        labels["version"],
    ] + [element_label_by_client_id.get(client_id, client_id) for client_id in element_client_ids]

    rows: list[list[str]] = []
    for response_row in responses:
        user = users_map.get(response_row.user_id)
        answers = answer_rows_by_response_id.get(response_row.response_id, {})
        row = [
            user.name if user else f"User {response_row.user_id}",
            user.email if user else "",
            str(_enum_value(response_row.status)),
            response_row.created_at.isoformat() if response_row.created_at else "",
            response_row.completed_at.isoformat() if response_row.completed_at else "",
            str(form_version),
        ]
        for client_id in element_client_ids:
            row.append(_stringify_export_value(answers.get(client_id), locale, list_sep))
        rows.append(row)
    return headers, rows


def _build_csv_export(headers: list[str], rows: list[list[str]]) -> bytes:
    out = io.StringIO(newline="")
    writer = csv.writer(out, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows(rows)
    return ("\ufeff" + out.getvalue()).encode("utf-8")


def _apply_xlsx_sheet_style(sheet: Any) -> None:
    from openpyxl.styles import Alignment, Font

    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_len = 10
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 80)


def _build_xlsx_export(
    *,
    locale: str,
    headers: list[str],
    rows: list[list[str]],
    summary_rows: list[list[str]],
    statistics_rows: list[list[str]],
) -> bytes:
    from openpyxl import Workbook

    names = EXPORT_SHEETS.get(locale, EXPORT_SHEETS["en"])
    workbook = Workbook()
    main = workbook.active
    main.title = names["responses"]
    main.append(headers)
    for row in rows:
        main.append(row)
    _apply_xlsx_sheet_style(main)

    ws_summary = workbook.create_sheet(names["summary"])
    for row in summary_rows:
        ws_summary.append(row)
    _apply_xlsx_sheet_style(ws_summary)

    ws_stats = workbook.create_sheet(names["statistics"])
    for row in statistics_rows:
        ws_stats.append(row)
    _apply_xlsx_sheet_style(ws_stats)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@router.get("/{form_id}/public", response_model=PublicFormDetailResponse)
async def get_public_form(
    form_id: int,
    key: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser | None = Depends(get_optional_user),
):
    form = await _get_published_form(db, form_id)
    form = await _resolve_latest_submitted_form(db, form)
    _check_form_access(form, key, current_user)
    responder = await _resolve_responder(db, current_user)

    attempts_remaining: int | None = None
    settings = form.settings_json if isinstance(form.settings_json, dict) else {}
    attempt_limit_type = settings.get("attemptLimitType", "unlimited")
    if attempt_limit_type == "limited":
        attempt_limit = settings.get("attemptLimit")
        if attempt_limit is not None and isinstance(attempt_limit, (int, float)):
            attempt_limit = int(attempt_limit)
            if attempt_limit > 0:
                # Считаем submitted всегда; cancelled — только если при отзыве правило уже действовало
                attempts_result = await db.execute(
                    select(Response)
                    .where(Response.form_id == form.form_id)
                    .where(Response.user_id == responder.user_id)
                    .where(
                        or_(
                            Response.status == "submitted",
                            and_(
                                Response.status == "cancelled",
                                Response.revoke_counts_as_attempt_at_revoke == True,
                            ),
                        )
                    )
                )
                attempts_used = len(attempts_result.scalars().all())
                attempts_remaining = max(0, attempt_limit - attempts_used)

    _increment_link_views(form)
    await db.flush()
    await db.refresh(form)

    detail = await build_form_detail_response(db, form)
    elements = [
        BuilderElementOut(**{**element.model_dump(), "correct_answer": None})
        for element in detail.elements
    ]

    return PublicFormDetailResponse(
        form_id=detail.form_id,
        title=detail.title,
        description=detail.description,
        settings_json=_public_settings(form),
        start_at=detail.start_at,
        end_at=detail.end_at,
        access_mode=detail.access_mode,
        status=detail.status,
        version=detail.version,
        prev_form_id=detail.prev_form_id,
        created_at=detail.created_at,
        updated_at=detail.updated_at,
        pages=detail.pages,
        elements=elements,
        conditions=detail.conditions,
        attempts_remaining=attempts_remaining,
    )


@router.get("/{form_id}/responses/draft", response_model=FormDraftResponse | None)
async def get_form_draft(
    form_id: int,
    key: str | None = Query(None),
    respondent_session_token: str | None = Query(None, alias="session_token"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser | None = Depends(get_optional_user),
):
    """Получить черновик ответа на форму (если есть)."""
    form = await _get_published_form(db, form_id)
    form = await _resolve_latest_submitted_form(db, form)
    _check_form_access(form, key, current_user)
    responder = await _resolve_responder(db, current_user)

    if current_user is not None:
        q = (
            select(Response)
            .where(Response.form_id == form.form_id)
            .where(Response.user_id == responder.user_id)
            .where(Response.status == "draft")
            .order_by(Response.created_at.desc())
            .limit(1)
        )
    else:
        if not respondent_session_token or not respondent_session_token.strip():
            return None
        token = respondent_session_token.strip()[:255]
        try:
            UUID(token)
        except (ValueError, TypeError):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="session_token must be a valid UUID",
            )
        q = (
            select(Response)
            .where(Response.form_id == form.form_id)
            .where(Response.respondent_session_token == token)
            .where(Response.status == "draft")
            .order_by(Response.created_at.desc())
            .limit(1)
        )
    result = await db.execute(q)
    draft = result.scalar_one_or_none()
    if not draft:
        return None

    answers_result = await db.execute(
        select(ResponseAnswer).where(ResponseAnswer.response_id == draft.response_id)
    )
    answer_rows = answers_result.scalars().all()
    elements_result = await db.execute(
        select(FormElement).where(FormElement.form_id == form.form_id)
    )
    elements = elements_result.scalars().all()
    by_element_id = {e.element_id: e for e in elements}
    element_client_map: dict[int, str] = {}
    for element in elements:
        settings = element.other_settings if isinstance(element.other_settings, dict) else {}
        client_id = settings.get("client_id")
        element_client_map[element.element_id] = (
            str(client_id) if client_id is not None else str(element.element_id)
        )

    answer_ids = [row.answer_id for row in answer_rows]
    files_by_answer_id: dict[int, list[dict[str, Any]]] = {}
    if answer_ids:
        files_result = await db.execute(
            select(UploadedFile)
            .where(UploadedFile.answer_id.in_(answer_ids))
            .where(UploadedFile.status != "deleted")
        )
        for file_row in files_result.scalars().all():
            if file_row.answer_id is None:
                continue
            files_by_answer_id.setdefault(file_row.answer_id, []).append(
                {
                    "file_id": file_row.file_id,
                    "name": file_row.name,
                    "mime_type": file_row.mime_type,
                    "size_bytes": file_row.size_bytes,
                    "url": _file_url(file_row.file_id, file_row.access_token),
                    "content_hash": file_row.content_hash,
                    "status": _enum_value(file_row.status),
                }
            )

    answers: dict[str, Any] = {}
    for answer_row in answer_rows:
        client_id = element_client_map.get(answer_row.element_id, str(answer_row.element_id))
        attachments = files_by_answer_id.get(answer_row.answer_id, [])
        value = _deserialize_answer_value(answer_row, attachments)
        answers[client_id] = value

    return FormDraftResponse(
        response_id=draft.response_id,
        answers=answers,
        respondent_session_token=getattr(draft, "respondent_session_token", None),
    )


@router.put("/{form_id}/responses/draft", response_model=FormDraftResponse)
async def save_form_draft(
    form_id: int,
    payload: FormDraftSaveRequest,
    key: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser | None = Depends(get_optional_user),
):
    """Сохранить черновик ответа на форму (создать или обновить)."""
    form = await _get_published_form(db, form_id)
    form = await _resolve_latest_submitted_form(db, form)
    _check_form_access(form, key, current_user)
    responder = await _resolve_responder(db, current_user)

    elements_result = await db.execute(
        select(FormElement).where(FormElement.form_id == form.form_id)
    )
    form_elements = elements_result.scalars().all()
    by_client_id: dict[str, FormElement] = {}
    for element in form_elements:
        settings = element.other_settings if isinstance(element.other_settings, dict) else {}
        client_id = settings.get("client_id")
        resolved_id = str(client_id) if client_id is not None else str(element.element_id)
        by_client_id[resolved_id] = element

    unknown_ids = [
        cid for cid in payload.answers.keys()
        if cid not in by_client_id
    ]
    if unknown_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown element id(s): {', '.join(unknown_ids[:5])}",
        )

    session_token: str | None = None
    if current_user is None:
        raw = payload.respondent_session_token
        session_token = raw.strip()[:255] if isinstance(raw, str) and raw.strip() else None
        if not session_token:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="respondent_session_token required for unauthenticated users",
            )

    if current_user is not None:
        draft_q = (
            select(Response)
            .where(Response.form_id == form.form_id)
            .where(Response.user_id == responder.user_id)
            .where(Response.status == "draft")
            .order_by(Response.created_at.desc())
            .limit(1)
        )
    else:
        draft_q = (
            select(Response)
            .where(Response.form_id == form.form_id)
            .where(Response.respondent_session_token == session_token)
            .where(Response.status == "draft")
            .order_by(Response.created_at.desc())
            .limit(1)
        )
    draft_result = await db.execute(draft_q)
    draft = draft_result.scalar_one_or_none()

    if draft is None:
        draft = Response(
            user_id=responder.user_id,
            form_id=form.form_id,
            status="draft",
            respondent_session_token=session_token,
        )
        db.add(draft)
        await db.flush()

    # Загружаем существующие ответы по response_id; перезаписываем по element_id, лишние удаляем
    existing_result = await db.execute(
        select(ResponseAnswer).where(ResponseAnswer.response_id == draft.response_id)
    )
    existing_by_element: dict[int, ResponseAnswer] = {
        row.element_id: row for row in existing_result.scalars().all()
    }

    payload_element_ids: set[int] = set()
    for client_id, value in payload.answers.items():
        element = by_client_id.get(client_id)
        if element is None:
            continue
        payload_element_ids.add(element.element_id)
        semantic_error = _validate_semantic_answer(element, value)
        if semantic_error is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid value for '{client_id}': {semantic_error}",
            )
        answer_row = existing_by_element.get(element.element_id)
        if answer_row is not None:
            _apply_answer_value(answer_row, value)
        else:
            answer_row = ResponseAnswer(
                response_id=draft.response_id,
                element_id=element.element_id,
            )
            _apply_answer_value(answer_row, value)
            db.add(answer_row)
            existing_by_element[element.element_id] = answer_row

    for eid, answer_row in list(existing_by_element.items()):
        if eid not in payload_element_ids:
            db.delete(answer_row)

    await db.flush()

    for client_id, value in payload.answers.items():
        element = by_client_id.get(client_id)
        if element is None:
            continue
        file_ids = _extract_file_ids(value)
        if file_ids:
            answer_row = existing_by_element.get(element.element_id)
            if answer_row is not None:
                await db.execute(
                    update(UploadedFile)
                    .where(UploadedFile.file_id.in_(file_ids))
                    .values(
                        answer_id=answer_row.answer_id,
                        status="temp",
                        expires_at=datetime.now(timezone.utc) + timedelta(hours=24),
                    )
                )
    await db.refresh(draft)

    answers_out: dict[str, Any] = {}
    for client_id, value in payload.answers.items():
        answers_out[client_id] = value

    return FormDraftResponse(
        response_id=draft.response_id,
        answers=answers_out,
        respondent_session_token=draft.respondent_session_token,
    )


@router.post(
    "/{form_id}/responses",
    response_model=FormSubmitAnswersResponse,
    status_code=status.HTTP_201_CREATED,
)
async def submit_form_response(
    form_id: int,
    payload: FormSubmitAnswersRequest,
    key: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser | None = Depends(get_optional_user),
):
    form = await _get_published_form(db, form_id)
    form = await _resolve_latest_submitted_form(db, form)
    _check_form_access(form, key, current_user)

    responder = await _resolve_responder(db, current_user)
    
    # Проверка лимита попыток
    settings = form.settings_json if isinstance(form.settings_json, dict) else {}
    attempt_limit_type = settings.get("attemptLimitType", "unlimited")
    
    if attempt_limit_type == "limited":
        attempt_limit = settings.get("attemptLimit")
        if attempt_limit is not None and isinstance(attempt_limit, (int, float)):
            attempt_limit = int(attempt_limit)
            if attempt_limit > 0:
                # Считаем submitted всегда; cancelled — только если при отзыве правило уже действовало
                attempts_result = await db.execute(
                    select(Response)
                    .where(Response.form_id == form.form_id)
                    .where(Response.user_id == responder.user_id)
                    .where(
                        or_(
                            Response.status == "submitted",
                            and_(
                                Response.status == "cancelled",
                                Response.revoke_counts_as_attempt_at_revoke == True,
                            ),
                        )
                    )
                )
                attempts_count = len(attempts_result.scalars().all())

                if attempts_count >= attempt_limit:
                    raise HTTPException(
                        status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                        detail=f"Attempt limit reached ({attempt_limit}). You have used all available attempts."
                    )
    
    elements_result = await db.execute(select(FormElement).where(FormElement.form_id == form.form_id))
    form_elements = elements_result.scalars().all()

    by_client_id: dict[str, FormElement] = {}
    for element in form_elements:
        settings = element.other_settings if isinstance(element.other_settings, dict) else {}
        client_id = settings.get("client_id")
        resolved_id = str(client_id) if client_id is not None else str(element.element_id)
        by_client_id[resolved_id] = element

    unknown_ids = [client_id for client_id in payload.answers.keys() if client_id not in by_client_id]
    if unknown_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown element id(s): {', '.join(unknown_ids[:5])}",
        )

    completed_at = datetime.now(timezone.utc)
    started_at = completed_at
    if payload.started_at is not None:
        candidate = payload.started_at
        if candidate.tzinfo is None:
            candidate = candidate.replace(tzinfo=timezone.utc)
        else:
            candidate = candidate.astimezone(timezone.utc)
        if candidate > completed_at:
            candidate = completed_at
        max_history = completed_at - timedelta(hours=24)
        if candidate < max_history:
            candidate = max_history
        started_at = candidate

    response_row: Response | None = None
    if payload.draft_response_id is not None:
        draft_result = await db.execute(
            select(Response)
            .where(Response.response_id == payload.draft_response_id)
            .where(Response.form_id == form.form_id)
            .where(Response.status == "draft")
        )
        candidate_draft = draft_result.scalar_one_or_none()
        if candidate_draft is not None:
            if current_user is not None:
                if candidate_draft.user_id == responder.user_id:
                    response_row = candidate_draft
            else:
                if candidate_draft.respondent_session_token:
                    response_row = candidate_draft

    if response_row is None:
        response_row = Response(
            user_id=responder.user_id,
            form_id=form.form_id,
            created_at=started_at,
            status="submitted",
            completed_at=completed_at,
        )
        db.add(response_row)
        await db.flush()
    else:
        response_row.status = "submitted"
        response_row.completed_at = completed_at
        created_at_val = response_row.created_at
        if created_at_val is not None and created_at_val.tzinfo is None:
            created_at_val = created_at_val.replace(tzinfo=timezone.utc)
        if response_row.created_at is None or (created_at_val is not None and created_at_val > started_at):
            response_row.created_at = started_at
        existing_answers_result = await db.execute(
            select(ResponseAnswer).where(ResponseAnswer.response_id == response_row.response_id)
        )
        for old_answer in existing_answers_result.scalars().all():
            db.delete(old_answer)

    answers_count = 0
    for client_id, value in payload.answers.items():
        element = by_client_id.get(client_id)
        if element is None:
            continue
        semantic_error = _validate_semantic_answer(element, value)
        if semantic_error is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid value for '{client_id}': {semantic_error}",
            )

        answer_row = ResponseAnswer(
            response_id=response_row.response_id,
            element_id=element.element_id,
        )
        _apply_answer_value(answer_row, value)
        db.add(answer_row)
        await db.flush()

        file_ids = _extract_file_ids(value)
        if file_ids:
            await db.execute(
                update(UploadedFile)
                .where(UploadedFile.file_id.in_(file_ids))
                .values(
                    answer_id=answer_row.answer_id,
                    status="submitted",
                    expires_at=None,
                )
            )
        answers_count += 1

    await db.flush()

    return FormSubmitAnswersResponse(
        response_id=response_row.response_id,
        submitted_at=response_row.completed_at or completed_at,
        answers_count=answers_count,
    )


@router.get("/{form_id}/responses", response_model=FormStoredResponsesResponse)
async def get_form_responses(
    form_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    form = await _ensure_editor_or_owner(db, form_id, current_user, allowed_roles=("editor", "participant"))

    responses_result = await db.execute(
        select(Response)
        .where(Response.form_id == form.form_id)
        .where(Response.status.in_(["submitted", "cancelled"]))
        .order_by(Response.completed_at.desc(), Response.created_at.desc())
    )
    responses = responses_result.scalars().all()
    if not responses:
        return FormStoredResponsesResponse(responses=[])

    response_ids = [item.response_id for item in responses]
    user_ids = list({item.user_id for item in responses})

    answers_result = await db.execute(
        select(ResponseAnswer).where(ResponseAnswer.response_id.in_(response_ids))
    )
    answer_rows = answers_result.scalars().all()

    elements_result = await db.execute(
        select(FormElement).where(FormElement.form_id == form.form_id)
    )
    elements = elements_result.scalars().all()
    element_client_map: dict[int, str] = {}
    for element in elements:
        settings = element.other_settings if isinstance(element.other_settings, dict) else {}
        client_id = settings.get("client_id")
        element_client_map[element.element_id] = str(client_id) if client_id is not None else str(element.element_id)

    users_result = await db.execute(select(AppUser).where(AppUser.user_id.in_(user_ids)))
    users_map = {user.user_id: user for user in users_result.scalars().all()}

    answer_ids = [row.answer_id for row in answer_rows]
    files_by_answer_id: dict[int, list[dict[str, Any]]] = {}
    if answer_ids:
        files_result = await db.execute(
            select(UploadedFile)
            .where(UploadedFile.answer_id.in_(answer_ids))
            .where(UploadedFile.status != "deleted")
        )
        for file_row in files_result.scalars().all():
            if file_row.answer_id is None:
                continue
            files_by_answer_id.setdefault(file_row.answer_id, []).append(
                {
                    "file_id": file_row.file_id,
                    "name": file_row.name,
                    "mime_type": file_row.mime_type,
                    "size_bytes": file_row.size_bytes,
                    "url": _file_url(file_row.file_id, file_row.access_token),
                    "content_hash": file_row.content_hash,
                    "status": _enum_value(file_row.status),
                }
            )

    answers_by_response_id: dict[int, dict[str, Any]] = {}
    for answer_row in answer_rows:
        client_id = element_client_map.get(answer_row.element_id, str(answer_row.element_id))
        attachments = files_by_answer_id.get(answer_row.answer_id, [])
        value = _deserialize_answer_value(answer_row, attachments)
        answers_by_response_id.setdefault(answer_row.response_id, {})[client_id] = value

    out: list[FormStoredResponse] = []
    form_version = form.version or 1
    for response_row in responses:
        user = users_map.get(response_row.user_id)
        out.append(
            FormStoredResponse(
                response_id=response_row.response_id,
                form_id=response_row.form_id,
                user_id=response_row.user_id,
                responder_name=user.name if user else f"User {response_row.user_id}",
                responder_email=user.email if user else None,
                status=_enum_value(response_row.status),
                created_at=response_row.created_at,
                completed_at=response_row.completed_at,
                version=form_version,
                answers=answers_by_response_id.get(response_row.response_id, {}),
            )
        )

    return FormStoredResponsesResponse(responses=out)


@router.get("/{form_id}/responses/export")
async def export_form_responses(
    form_id: int,
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    locale: str = Query("en"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    export_locale = "ru" if locale.lower().startswith("ru") else "en"
    form = await _ensure_editor_or_owner(db, form_id, current_user, allowed_roles=("editor", "participant"))

    responses_result = await db.execute(
        select(Response)
        .where(Response.form_id == form.form_id)
        .where(Response.status.in_(["submitted", "cancelled"]))
        .order_by(Response.response_id.asc())
    )
    responses = responses_result.scalars().all()

    response_ids = [item.response_id for item in responses]
    user_ids = list({item.user_id for item in responses})

    answers_result = await db.execute(
        select(ResponseAnswer).where(ResponseAnswer.response_id.in_(response_ids))
    ) if response_ids else None
    answer_rows = answers_result.scalars().all() if answers_result is not None else []

    elements_result = await db.execute(
        select(FormElement).where(FormElement.form_id == form.form_id)
    )
    elements = [
        item
        for item in elements_result.scalars().all()
        if _enum_value(item.widget) not in ("heading", "static_text")
    ]
    pages_result = await db.execute(
        select(FormPage).where(FormPage.form_id == form.form_id)
    )
    page_index_by_id = {
        int(page.page_id): int(page.page_index)
        for page in pages_result.scalars().all()
    }

    users_result = await db.execute(select(AppUser).where(AppUser.user_id.in_(user_ids))) if user_ids else None
    users_map = {user.user_id: user for user in (users_result.scalars().all() if users_result is not None else [])}

    answer_ids = [row.answer_id for row in answer_rows]
    files_by_answer_id: dict[int, list[dict[str, Any]]] = {}
    if answer_ids:
        files_result = await db.execute(
            select(UploadedFile)
            .where(UploadedFile.answer_id.in_(answer_ids))
            .where(UploadedFile.status != "deleted")
        )
        for file_row in files_result.scalars().all():
            if file_row.answer_id is None:
                continue
            files_by_answer_id.setdefault(file_row.answer_id, []).append(
                {
                    "file_id": file_row.file_id,
                    "name": file_row.name,
                    "mime_type": file_row.mime_type,
                    "size_bytes": file_row.size_bytes,
                    "url": _file_url(file_row.file_id, file_row.access_token),
                    "content_hash": file_row.content_hash,
                    "status": _enum_value(file_row.status),
                }
            )

    sorted_elements = _sorted_form_elements(elements, page_index_by_id)
    answer_rows_by_response_id = _build_answer_map_by_response(
        answer_rows, files_by_answer_id, sorted_elements
    )

    export_at = datetime.now(timezone.utc)
    date_part = export_at.strftime("%d-%m-%Y")
    time_part = export_at.strftime("%H:%M")

    base_title = _sanitize_filename_part(form.title or "")
    if not base_title:
        base_title = f"form-{form.form_id}"
    filename_base = f"{base_title}-{date_part}-{time_part}"
    ascii_prefix = f"form-{form.form_id}-{date_part}-{time_part}"

    if format == "csv":
        headers, rows = _build_export_rows(
            form_version=form.version or 1,
            responses=responses,
            answer_rows_by_response_id=answer_rows_by_response_id,
            users_map=users_map,
            sorted_elements=sorted_elements,
            locale=export_locale,
            list_sep=EXPORT_LIST_SEP_CSV,
        )
        content = _build_csv_export(headers, rows)
        media_type = "text/csv; charset=utf-8"
        filename = f"{filename_base}.csv"
        ascii_filename = f"{ascii_prefix}.csv"
    else:
        headers, rows = _build_export_rows(
            form_version=form.version or 1,
            responses=responses,
            answer_rows_by_response_id=answer_rows_by_response_id,
            users_map=users_map,
            sorted_elements=sorted_elements,
            locale=export_locale,
            list_sep="\n",
        )
        summary_rows = _build_summary_export_rows(
            locale=export_locale,
            form_title=form.title or "",
            form_version=form.version or 1,
            responses=responses,
            export_at=export_at,
        )
        statistics_rows = _build_statistics_export_rows(
            locale=export_locale,
            sorted_elements=sorted_elements,
            answer_rows_by_response_id=answer_rows_by_response_id,
            responses=responses,
        )
        content = _build_xlsx_export(
            locale=export_locale,
            headers=headers,
            rows=rows,
            summary_rows=summary_rows,
            statistics_rows=statistics_rows,
        )
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename_base}.xlsx"
        ascii_filename = f"{ascii_prefix}.xlsx"

    quoted_name = quote(filename)
    content_disposition = f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{quoted_name}"
    return FastApiResponse(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": content_disposition},
    )
