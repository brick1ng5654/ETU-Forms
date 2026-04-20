from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any

from app.models import AppUser, FormElement, Response

EXPORT_HEADERS = {
    "ru": {
        "responder_name": "Респондент",
        "responder_email": "Email",
        "status": "Статус",
        "version": "Версия формы",
        "other": "Другое",
    },
    "en": {
        "responder_name": "Responder",
        "responder_email": "Email",
        "status": "Status",
        "version": "Form version",
        "other": "Other",
    },
}

EXPORT_LIST_SEP = "\n"

EXPORT_SHEETS = {
    "ru": {"responses": "Ответы", "summary": "Итоги", "statistics": "Статистика"},
    "en": {"responses": "Responses", "summary": "Summary", "statistics": "Statistics"},
}

EXPORT_SUMMARY = {
    "ru": {
        "metric": "Показатель",
        "value": "Значение",
        "form_title": "Название формы",
        "exported_at": "Дата и время выгрузки",
        "responses_in_export": "Ответов в выгрузке",
        "status_submitted": "Отправлено",
        "status_cancelled": "Отменено",
        "form_version": "Версия формы",
    },
    "en": {
        "metric": "Metric",
        "value": "Value",
        "form_title": "Form title",
        "exported_at": "Export date and time",
        "responses_in_export": "Responses in export",
        "status_submitted": "Submitted",
        "status_cancelled": "Cancelled",
        "form_version": "Form version",
    },
}

EXPORT_STATS = {
    "ru": {
        "dimension": "Показатель",
        "value_col": "Значение",
        "nonempty": "Ответов с данными по полю",
        "min": "Минимум",
        "max": "Максимум",
        "average": "Среднее",
        "distribution": "Распределение",
    },
    "en": {
        "dimension": "Metric",
        "value_col": "Value",
        "nonempty": "Non-empty answers for field",
        "min": "Minimum",
        "max": "Maximum",
        "average": "Average",
        "distribution": "Distribution",
    },
}

EXPORT_COMPLEX_FIELD_LABELS = {
    "ru": {
        "lastName": "Фамилия",
        "firstName": "Имя",
        "patronymic": "Отчество",
        "gender": "Пол",
        "birthDate": "Дата рождения",
        "seriesNumber": "Серия и номер",
        "issuedBy": "Кем выдан",
        "issueDate": "Дата выдачи",
        "departmentCode": "Код подразделения",
        "birthPlace": "Место рождения",
    },
    "en": {
        "lastName": "Last name",
        "firstName": "First name",
        "patronymic": "Middle name",
        "gender": "Gender",
        "birthDate": "Date of birth",
        "seriesNumber": "Series and number",
        "issuedBy": "Issued by",
        "issueDate": "Issue date",
        "departmentCode": "Department code",
        "birthPlace": "Place of birth",
    },
}

EXPORT_COMPLEX_FIELD_ORDER = (
    "lastName",
    "firstName",
    "patronymic",
    "gender",
    "birthDate",
    "seriesNumber",
    "issuedBy",
    "issueDate",
    "departmentCode",
    "birthPlace",
)

EXPORT_GENDER_LABELS = {
    "ru": {"male": "Мужской", "female": "Женский"},
    "en": {"male": "Male", "female": "Female"},
}

_XLSX_CELL_MAX_LEN = 32767
_ILLEGAL_XLSX_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _enum_value(x: Any) -> Any:
    return x.value if hasattr(x, "value") else x


def _extract_element_client_id(element: FormElement) -> str:
    settings = element.other_settings if isinstance(element.other_settings, dict) else {}
    client_id = settings.get("client_id")
    return str(client_id) if client_id is not None else str(element.element_id)


def _extract_element_settings(element: FormElement) -> dict[str, Any]:
    return element.other_settings if isinstance(element.other_settings, dict) else {}


def _is_repeatable_block_element(element: FormElement) -> bool:
    widget = _enum_value(element.widget)
    if widget == "repeatable_block":
        return True
    settings = _extract_element_settings(element)
    return widget == "text_input" and settings.get("repeatableBlock") is True


def _extract_repeatable_base_name(element: FormElement, locale: str) -> str:
    settings = _extract_element_settings(element)
    base = str(settings.get("instanceNameBase") or "").strip()
    if base:
        return base
    label = str(element.label or "").strip()
    if label:
        return label
    return "Блок" if locale == "ru" else "Block"


def _extract_repeatable_max_count(element: FormElement) -> int | None:
    settings = _extract_element_settings(element)
    raw = settings.get("maxCount")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def _split_repeatable_elements(
    sorted_elements: list[FormElement],
) -> tuple[list[FormElement], list[FormElement], dict[str, list[FormElement]]]:
    repeatable_blocks = [element for element in sorted_elements if _is_repeatable_block_element(element)]
    repeatable_block_ids = {_extract_element_client_id(element) for element in repeatable_blocks}
    children_by_block_id: dict[str, list[FormElement]] = {block_id: [] for block_id in repeatable_block_ids}
    regular_elements: list[FormElement] = []

    for element in sorted_elements:
        client_id = _extract_element_client_id(element)
        if client_id in repeatable_block_ids:
            continue
        settings = _extract_element_settings(element)
        parent_block_id = str(settings.get("parentBlockId") or "").strip()
        if parent_block_id and parent_block_id in repeatable_block_ids:
            children_by_block_id[parent_block_id].append(element)
            continue
        regular_elements.append(element)

    return regular_elements, repeatable_blocks, children_by_block_id


def _prepare_xlsx_cell_value(value: Any) -> str:
    if value is None:
        text = ""
    else:
        text = str(value)
    text = _ILLEGAL_XLSX_CTRL.sub("", text)
    if len(text) > _XLSX_CELL_MAX_LEN:
        text = text[: _XLSX_CELL_MAX_LEN - 1] + "…"
    return text


def _extract_labeled_dict_parts(value: dict[str, Any], locale: str, list_sep: str) -> list[str]:
    labels = EXPORT_COMPLEX_FIELD_LABELS.get(locale, EXPORT_COMPLEX_FIELD_LABELS["en"])
    ordered_keys = [key for key in EXPORT_COMPLEX_FIELD_ORDER if key in value]
    remaining_keys = [key for key in value.keys() if key not in EXPORT_COMPLEX_FIELD_ORDER]
    out: list[str] = []

    for key in [*ordered_keys, *remaining_keys]:
        raw = value.get(key)
        if raw is None:
            continue
        if isinstance(raw, str):
            text = raw.strip()
        elif key == "gender":
            normalized_gender = str(raw).strip().lower()
            text = EXPORT_GENDER_LABELS.get(locale, EXPORT_GENDER_LABELS["en"]).get(
                normalized_gender, str(raw)
            )
        elif isinstance(raw, dict):
            text = json.dumps(raw, ensure_ascii=False, default=str)
        else:
            text = _stringify_export_value(raw, locale, list_sep).strip()
        if not text:
            continue
        label = labels.get(key, key)
        out.append(f"{label}: {text}")
    return out


def _extract_export_answer_parts(value: Any, locale: str, list_sep: str) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            text = _stringify_export_value(item, locale, list_sep)
            if text:
                out.append(text)
        return out
    if isinstance(value, dict):
        if "attachments" in value and isinstance(value.get("attachments"), list):
            return _extract_export_answer_parts(value.get("attachments"), locale, list_sep)
        if "name" in value and isinstance(value.get("name"), str):
            name = str(value.get("name")).strip()
            return [name] if name else []
        if "url" in value and isinstance(value.get("url"), str):
            url = str(value.get("url")).strip()
            return [url] if url else []
        if "selected" in value or "otherSelected" in value:
            selected = value.get("selected")
            selected_parts: list[str] = []
            if isinstance(selected, list):
                selected_parts.extend([str(item) for item in selected if str(item).strip()])
            elif selected is not None and str(selected).strip():
                selected_parts.append(str(selected))
            if value.get("otherSelected"):
                other_text = str(value.get("otherText") or "").strip()
                other_label = EXPORT_HEADERS.get(locale, EXPORT_HEADERS["en"])["other"]
                selected_parts.append(f"{other_label}: {other_text}" if other_text else other_label)
            return selected_parts
        if "date" in value or "time" in value:
            date_part = str(value.get("date") or "").strip()
            time_part = str(value.get("time") or "").strip()
            joined = " ".join([part for part in [date_part, time_part] if part]).strip()
            return [joined] if joined else []
        labeled_parts = _extract_labeled_dict_parts(value, locale, list_sep)
        if labeled_parts:
            return labeled_parts
        text = json.dumps(value, ensure_ascii=False, default=str).strip()
        return [text] if text and text != "{}" else []
    text = _stringify_export_value(value, locale, list_sep).strip()
    return [text] if text else []


def _stringify_export_value(value: Any, locale: str, list_sep: str) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return list_sep.join(_extract_export_answer_parts(value, locale, list_sep))
    if isinstance(value, dict):
        parts = _extract_export_answer_parts(value, locale, list_sep)
        if parts:
            return list_sep.join(parts)
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _is_nonempty_export_answer(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, list):
        return len(value) > 0
    if isinstance(value, bool):
        return True
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, dict):
        return True
    return True


def _collect_numeric_values_for_stats(values: list[Any]) -> list[float] | None:
    nums: list[float] = []
    for raw in values:
        if raw is None:
            continue
        if isinstance(raw, bool):
            return None
        if isinstance(raw, (int, float)):
            nums.append(float(raw))
            continue
        if isinstance(raw, str):
            s = raw.strip().replace(",", ".")
            if not s:
                continue
            try:
                nums.append(float(s))
            except ValueError:
                return None
            continue
        return None
    return nums if nums else None


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", value or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip().strip(".")
    return cleaned


def build_export_rows(
    *,
    form_version: int,
    responses: list[Response],
    answer_rows_by_response_id: dict[int, dict[str, Any]],
    users_map: dict[int, AppUser],
    sorted_elements: list[FormElement],
    locale: str,
    list_sep: str,
) -> tuple[list[str], list[list[str]]]:
    regular_elements, repeatable_blocks, repeatable_children_by_block = _split_repeatable_elements(sorted_elements)
    regular_element_client_ids = [_extract_element_client_id(element) for element in regular_elements]
    regular_label_by_client_id = {
        _extract_element_client_id(element): (element.label or f"Field {element.element_id}").strip()
        or f"Field {element.element_id}"
        for element in regular_elements
    }

    repeatable_specs: list[dict[str, Any]] = []
    for block in repeatable_blocks:
        block_client_id = _extract_element_client_id(block)
        children = repeatable_children_by_block.get(block_client_id, [])
        child_specs = [
            {
                "client_id": _extract_element_client_id(child),
                "label": (child.label or f"Field {child.element_id}").strip() or f"Field {child.element_id}",
            }
            for child in children
        ]
        observed_instances = 0
        for response_answers in answer_rows_by_response_id.values():
            raw = response_answers.get(block_client_id)
            if isinstance(raw, list):
                observed_instances = max(observed_instances, len(raw))
        configured_max_count = _extract_repeatable_max_count(block)
        instances_count = max(
            configured_max_count or 0,
            observed_instances,
        )
        if instances_count <= 0:
            instances_count = 1
        repeatable_specs.append(
            {
                "client_id": block_client_id,
                "base_name": _extract_repeatable_base_name(block, locale),
                "instances_count": instances_count,
                "children": child_specs,
            }
        )

    labels = EXPORT_HEADERS.get(locale, EXPORT_HEADERS["en"])
    headers = [
        labels["responder_name"],
        labels["responder_email"],
        labels["status"],
        labels["version"],
    ] + [regular_label_by_client_id.get(client_id, client_id) for client_id in regular_element_client_ids]
    for spec in repeatable_specs:
        base_name = str(spec["base_name"])
        for index in range(1, int(spec["instances_count"]) + 1):
            headers.append(f"{base_name} {index}")

    rows: list[list[str]] = []
    for response_row in responses:
        user = users_map.get(response_row.user_id)
        answers = answer_rows_by_response_id.get(response_row.response_id, {})
        row = [
            (user.name or f"User {response_row.user_id}") if user else f"User {response_row.user_id}",
            (user.email or "") if user else "",
            str(_enum_value(response_row.status)),
            str(form_version),
        ]
        for client_id in regular_element_client_ids:
            row.append(_stringify_export_value(answers.get(client_id), locale, list_sep))

        for spec in repeatable_specs:
            block_client_id = str(spec["client_id"])
            block_children = spec["children"]
            raw_instances = answers.get(block_client_id)
            instances = raw_instances if isinstance(raw_instances, list) else []
            instances_count = int(spec["instances_count"])
            for instance_index in range(instances_count):
                if instance_index >= len(instances) or not isinstance(instances[instance_index], dict):
                    row.append("")
                    continue
                instance_data = instances[instance_index]
                if not block_children:
                    row.append(_stringify_export_value(instance_data, locale, list_sep))
                    continue
                parts: list[str] = []
                for child_spec in block_children:
                    child_client_id = str(child_spec["client_id"])
                    child_label = str(child_spec["label"])
                    child_raw_value = instance_data.get(child_client_id)
                    if isinstance(child_raw_value, dict):
                        nested_parts = _extract_labeled_dict_parts(child_raw_value, locale, list_sep)
                        if nested_parts:
                            parts.extend(nested_parts)
                            continue
                    child_value = _stringify_export_value(child_raw_value, locale, list_sep).strip()
                    parts.append(f"{child_label}: {child_value}")
                row.append(list_sep.join(parts))
        rows.append(row)
    return headers, rows


def build_summary_export_rows(
    *,
    locale: str,
    form_title: str,
    form_version: int,
    responses: list[Response],
    export_at: datetime,
) -> list[list[str]]:
    labels = EXPORT_SUMMARY.get(locale, EXPORT_SUMMARY["en"])
    submitted = sum(1 for r in responses if _enum_value(r.status) == "submitted")
    cancelled = sum(1 for r in responses if _enum_value(r.status) == "cancelled")
    ts = export_at.isoformat(timespec="seconds")
    return [
        [labels["metric"], labels["value"]],
        [labels["form_title"], form_title],
        [labels["form_version"], str(form_version)],
        [labels["exported_at"], ts],
        [labels["responses_in_export"], str(len(responses))],
        [labels["status_submitted"], str(submitted)],
        [labels["status_cancelled"], str(cancelled)],
    ]


def build_statistics_export_rows(
    *,
    locale: str,
    sorted_elements: list[FormElement],
    answer_rows_by_response_id: dict[int, dict[str, Any]],
    responses: list[Response],
) -> list[list[str]]:
    stats_labels = EXPORT_STATS.get(locale, EXPORT_STATS["en"])
    rows: list[list[str]] = [
        [stats_labels["dimension"], stats_labels["value_col"]],
    ]
    list_sep = EXPORT_LIST_SEP
    regular_elements, repeatable_blocks, _ = _split_repeatable_elements(sorted_elements)
    elements_for_stats = [*regular_elements, *repeatable_blocks]

    for element in elements_for_stats:
        widget = _enum_value(element.widget)
        if widget in ("heading", "static_text"):
            continue
        client_id = _extract_element_client_id(element)
        label = (element.label or "").strip() or f"Field {element.element_id}"
        rows.append([f"— {label} —", ""])

        per_response: list[Any] = []
        for response_row in responses:
            raw = answer_rows_by_response_id.get(response_row.response_id, {}).get(client_id)
            per_response.append(raw)

        if _is_repeatable_block_element(element):
            counts: list[int] = []
            for raw in per_response:
                if isinstance(raw, list):
                    counts.append(len(raw))
                elif raw is None:
                    counts.append(0)
                else:
                    counts.append(1)
            nonempty_count = sum(1 for count in counts if count > 0)
            rows.append([stats_labels["nonempty"], str(nonempty_count)])
            if counts:
                rows.append([stats_labels["min"], str(min(counts))])
                rows.append([stats_labels["max"], str(max(counts))])
                rows.append([stats_labels["average"], str(round(sum(counts) / len(counts), 6))])
            rows.append(["", ""])
            continue

        nonempty = [v for v in per_response if _is_nonempty_export_answer(v)]
        rows.append([stats_labels["nonempty"], str(len(nonempty))])

        if not nonempty:
            rows.append(["", ""])
            continue

        nums = _collect_numeric_values_for_stats(nonempty)
        if (
            nums is not None
            and len(nums) == len(nonempty)
            and widget in ("number_input", "rating")
        ):
            rows.append([stats_labels["min"], str(min(nums))])
            rows.append([stats_labels["max"], str(max(nums))])
            rows.append([stats_labels["average"], str(round(sum(nums) / len(nums), 6))])
            rows.append(["", ""])
            continue

        counter: dict[str, int] = {}
        for v in nonempty:
            for part in _extract_export_answer_parts(v, locale, list_sep):
                p = part.strip()
                if not p:
                    continue
                counter[p] = counter.get(p, 0) + 1
        if not counter:
            rows.append(["", ""])
            continue
        rows.append([stats_labels["distribution"], ""])
        for key in sorted(counter.keys(), key=lambda s: (-counter[s], s)):
            rows.append([key, str(counter[key])])
        rows.append(["", ""])

    while rows and rows[-1] == ["", ""]:
        rows.pop()
    return rows


def _apply_xlsx_sheet_style(sheet: Any) -> None:
    from openpyxl.styles import Alignment, Font

    if not sheet.max_row or not sheet.max_column:
        return
    sheet.freeze_panes = "A2" if sheet.max_row > 1 else None
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_len = 10
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 80)


def build_xlsx_export(
    *,
    locale: str,
    headers: list[str],
    rows: list[list[str]],
    summary_rows: list[list[str]],
    statistics_rows: list[list[str]],
) -> bytes:
    from openpyxl import Workbook

    names = EXPORT_SHEETS.get(locale, EXPORT_SHEETS["en"])
    workbook = Workbook()
    main = workbook.active
    main.title = names["responses"]

    def _append_prepared(ws: Any, row: list[Any]) -> None:
        ws.append([_prepare_xlsx_cell_value(v) for v in row])

    _append_prepared(main, headers)
    for row in rows:
        _append_prepared(main, row)
    _apply_xlsx_sheet_style(main)

    ws_summary = workbook.create_sheet(names["summary"])
    for row in summary_rows:
        _append_prepared(ws_summary, row)
    _apply_xlsx_sheet_style(ws_summary)

    ws_stats = workbook.create_sheet(names["statistics"])
    for row in statistics_rows:
        _append_prepared(ws_stats, row)
    _apply_xlsx_sheet_style(ws_stats)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
